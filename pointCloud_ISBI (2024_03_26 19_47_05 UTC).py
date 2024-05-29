# -*- encoding: utf-8 -*-
import os
import platform
import sys
import argparse
import numpy as np
import SimpleITK as sitk
import skimage.morphology as skm
from tqdm import tqdm
import pandas as pd
from openpyxl import Workbook, load_workbook

if os.path.abspath("..") not in sys.path:
    sys.path.insert(0, os.path.abspath(".."))

from astra.utils.data_utils import (
    read_data,
    pre_processing,
    test_time_augmentation,
    copy_sitk_imageinfo,
)
from astra.model.model import Model
from astra.training.network_trainer import *

PERT_SIZE = 3
PERT_TYPE = "D"


def find_boundary_points(volume):
    """
    Find points on the boundary of a region of interest.
    These points will then be used to create perturbations.
    """
    ball = skm.ball(1)
    volume_larger = skm.binary_dilation(volume[0, :, :, :], ball)
    boundary_volume = volume_larger - volume[0, :, :, :]
    points = np.nonzero(boundary_volume)
    out_points = []

    # Choose 10 here to sub-sample the surface. Need to think of a better way to do this.
    for idx in range(0, len(points[0]), 2):
        x = points[0][idx]
        y = points[1][idx]
        z = points[2][idx]
        out_points.append([x, y, z])
    return out_points


def find_boundary_points_CTV(volume):
    """
    Find points on the boundary of a region of interest.
    These points will then be used to create perturbations.
    """
    ball = skm.ball(1)
    volume_smaller = skm.binary_erosion(volume[0, :, :, :], ball)
    boundary_rim = volume[0, :, :, :] - volume_smaller
    points = np.nonzero(boundary_rim)
    out_points = []

    # Choose 10 here to sub-sample the surface. Need to think of a better way to do this.
    for idx in range(0, len(points[0]), 2):
        x = points[0][idx]
        y = points[1][idx]
        z = points[2][idx]
        out_points.append([x, y, z])
    return out_points


def dilate_at(volume, point):
    """
    Dilate the binary volume 'volume' at the point specified bt point.
    """
    # print(PERT_SIZE)
    ball = skm.ball(PERT_SIZE)
    # print(str(np.count_nonzero(ball)))
    # print(str(np.count_nonzero(volume)))
    point_vol = np.zeros(volume[0, :, :, :].shape, dtype=np.uint8)
    point_vol[point[0], point[1], point[2]] = 1
    volume_out = skm.binary_dilation(point_vol, ball).astype(np.uint8)
    volume_out += volume[0, :, :, :].astype(np.uint8)
    volume_out[volume_out >= 1] = 1
    volume_out = volume_out[np.newaxis, :, :, :]
    # print(str(np.count_nonzero(volume_out)))
    return volume_out


def erode_at(volume, point):
    """
    Erode the binary volume 'volume' at the point specified bt point.
    """
    ball = skm.ball(PERT_SIZE)
    # print(str(np.count_nonzero(ball)))
    # print(str(np.count_nonzero(volume)))
    point_vol = np.zeros(volume[0, :, :, :].shape, dtype=np.uint8)
    point_vol[point[0], point[1], point[2]] = 1
    volume_out = skm.binary_dilation(point_vol, ball).astype(np.uint8)
    volume_out = volume[0, :, :, :].astype(np.uint8) - volume_out
    volume_out[volume_out >= 2] = 0
    volume_out = volume_out[np.newaxis, :, :, :]
    # print(str(np.count_nonzero(volume_out)))
    return volume_out


def inference_with_perturbation(trainer, list_patient_dirs, save_path, do_TTA=True):
    """
    This function helps create perturbations in the OAR and the Target, and then evaluates the dose.
    """
    sys = platform.system()

    os.makedirs(save_path, exist_ok=True)

    with torch.no_grad():
        trainer.setting.network.eval()
        for patient_dir in list_patient_dirs:
            if sys == 'Windows':
                patient_id = patient_dir.split("\\")[-1]
            else:
                patient_id = patient_dir.split("/")[-1]

            nameFile = 'PointCloud_' + patient_id + 'OverCont.xlsx'
            sheetName = str(patient_id)

            try:
                wb = load_workbook(nameFile)
            except FileNotFoundError:
                wb = Workbook()


            ws = wb.active

            if sheetName in wb.sheetnames:
                checkSheet = 'Sheet good'
            else:
                wb.create_sheet(sheetName)

            sheet = wb[sheetName]

            # sheet = wb.active if sheetName in wb.sheetnames else wb.create_sheet(title=sheetName)

            header = ['organ', 'x0', 'y0', 'z0', 'x1', 'y1', 'z1', 'D1', 'x2', 'y2', 'z2', 'D2', 'x3', 'y3', 'z3', 'D3', 'x4',
                      'y4', 'z4', 'D4', 'x5', 'y5', 'z5', 'D5', 'x6', 'y6', 'z6', 'D6', 'x7', 'y7', 'z7', 'D7', 'x8',
                      'y8', 'z8', 'D8', 'x9', 'y9', 'z9', 'D9', 'x10', 'y10', 'z10', 'D10', 'x11', 'y11', 'z11', 'D11',
                      'x12', 'y12', 'z12', 'D12', 'x13', 'y13', 'z13', 'D13', 'x14', 'y14', 'z14', 'D14', 'x15', 'y15',
                      'z15', 'D15', 'x16', 'y16', 'z16', 'D16', 'x17', 'y17', 'z17', 'D17', 'x18', 'y18', 'z18', 'D18',
                      'x19', 'y19', 'z19', 'D19', 'x20', 'y20', 'z20', 'D20']

            sheet.append(header)

            wb.save(nameFile)

            dict_images = read_data(patient_dir)

            list_images = pre_processing(dict_images)

            input_ = list_images[0]
            possible_dose_mask = list_images[1]

            # Test-time augmentation
            if do_TTA:
                TTA_mode = [[], ["Z"], ["W"], ["Z", "W"]]
            else:
                TTA_mode = [[]]
            prediction = test_time_augmentation(trainer, input_, TTA_mode)

            # Pose-processing
            prediction[
                np.logical_or(possible_dose_mask[0, :, :, :] < 1, prediction < 0)
            ] = 0
            gt_prediction = 70.0 * prediction

            templete_nii = sitk.ReadImage(patient_dir + "/Dose_Mask.nii.gz")
            prediction_nii = sitk.GetImageFromArray(gt_prediction)
            prediction_nii = copy_sitk_imageinfo(templete_nii, prediction_nii)

            if sys == 'Windows':
                os.makedirs(save_path + "\\" + patient_id, exist_ok=True)
                sitk.WriteImage(
                    prediction_nii,
                    save_path + "\\" + patient_id + "/Dose_gt.nii.gz",
                )
            else:
                os.makedirs(save_path + "/" + patient_id, exist_ok=True)
                sitk.WriteImage(
                    prediction_nii,
                    save_path + "/" + patient_id + "/Dose_gt.nii.gz",
                )

            list_target = ["Target"]
            list_oar_names = ["BrainStem", "Hippocampus_L", "Hippocampus_R", "Eye_L", "Eye_R", "Chiasm", "OpticNerve_L",
                              "OpticNerve_R"]  # "Cochlea_L", "Cochlea_R", "LacrimalGland_L", "LacrimalGland_R", "Pituitary"]

            for organ in list_target:

                print("Working on: ", organ.split("_")[0])

                # perturb_prediction = {}
                perturb_prediction_max = {}
                perturb_prediction_mean = {}
                perturb_prediction_dmax = {}
                perturb_prediction_dmean = {}
                perturb_prediction_hr_hl_thresh = {}

                perturb_point = {}
                perturb_pointcorrMax = {}

                perturb_point = np.zeros_like(gt_prediction)


                prediction_tv = np.zeros_like(gt_prediction)
                # perturb_prediction[organ] = np.zeros_like(gt_prediction)
                perturb_prediction_max[organ] = np.zeros_like(gt_prediction)
                perturb_prediction_mean[organ] = np.zeros_like(gt_prediction)
                perturb_prediction_dmax[organ] = np.zeros_like(gt_prediction)
                perturb_prediction_dmean[organ] = np.zeros_like(gt_prediction)

                for oar in list_oar_names:
                    # perturb_prediction[oar] = np.zeros_like(gt_prediction)
                    perturb_prediction_max[oar] = np.zeros_like(gt_prediction)
                    perturb_prediction_mean[oar] = np.zeros_like(gt_prediction)
                    perturb_prediction_dmax[oar] = np.zeros_like(gt_prediction)
                    perturb_prediction_dmean[oar] = np.zeros_like(gt_prediction)

                    if (oar == "Hippocampus_L") or (oar == "Hippocampus_R"):
                        perturb_prediction_hr_hl_thresh[oar] = np.zeros_like(gt_prediction)

                    perturb_pointcorrMax[oar] = np.zeros_like(gt_prediction)



                prediction_tv += np.multiply(gt_prediction, dict_images[organ][0, :, :, :])

                templete_nii = sitk.ReadImage(patient_dir + "/Dose_Mask.nii.gz")
                prediction_nii = sitk.GetImageFromArray(prediction_tv)
                prediction_nii = copy_sitk_imageinfo(templete_nii, prediction_nii)
                if sys == 'Windows':
                    os.makedirs(save_path + "\\" + patient_id, exist_ok=True)
                    sitk.WriteImage(
                        prediction_nii,
                        save_path + "\\" + patient_id + "/Prediction_NoPert" + organ + ".nii.gz",
                    )
                else:
                    os.makedirs(save_path + "/" + patient_id, exist_ok=True)
                    sitk.WriteImage(
                        prediction_nii,
                        save_path + "/" + patient_id + "/Prediction_NoPert" + organ + ".nii.gz",
                    )

                ### Use this to get boundary on the CTV
                point_set = find_boundary_points_CTV(dict_images[organ])


                print("\n Points on surface: ", len(point_set))

                # store mask of tv
                og_tv = dict_images[organ]

                short_ptSet = point_set[1:300:10]

                # At this stage, do perturbation on the organ boundary.
                for point in tqdm(short_ptSet):
                    # reset tv after perturbation, single perturbatione
                    dict_images[organ] = og_tv

                    try:
                        wb = load_workbook(nameFile)
                    except FileNotFoundError:
                        wb = Workbook()

                    sheetName = patient_id

                    ws = wb.active

                    if sheetName in wb.sheetnames:
                        checkSheet = 'Sheet good'
                    else:
                        wb.create_sheet(sheetName)

                    sheet = wb[sheetName]

                    if PERT_TYPE == "E":
                        dict_images[organ] = erode_at(dict_images[organ], point)
                    elif PERT_TYPE == "D":
                        dict_images[organ] = dilate_at(dict_images[organ], point)
                    else:
                        print("Not allowed argument.")

                    list_images = pre_processing(dict_images)

                    input_ = list_images[0]
                    possible_dose_mask = list_images[1]

                    # Test-time augmentation
                    if do_TTA:
                        TTA_mode = [[], ["Z"], ["W"], ["Z", "W"]]
                    else:
                        TTA_mode = [[]]
                    prediction = test_time_augmentation(trainer, input_, TTA_mode)

                    # Pose-processing
                    prediction[
                        np.logical_or(
                            possible_dose_mask[0, :, :, :] < 1, prediction < 0
                        )
                    ] = 0
                    # rescale and get gray (Gy)
                    prediction = 70.0 * prediction

                    temp_val = 1.00E-07

                    perturb_point[point[0], point[1], point[2]] = 1

                    prediction_pertPoint_nii = sitk.GetImageFromArray(perturb_point)

                    if sys == 'Windows':
                        os.makedirs(save_path + "\\" + patient_id, exist_ok=True)
                        sitk.WriteImage(
                            prediction_pertPoint_nii,
                            save_path + "\\" + patient_id + "/Perturbed_TV_PertPoint_" + str(point[0]) + "_" + str(
                                point[1]) + "_" + str(point[2]) + ".nii.gz",
                        )
                    else:
                        os.makedirs(save_path + "/" + patient_id, exist_ok=True)
                        sitk.WriteImage(
                            prediction_pertPoint_nii,
                            save_path + "/" + patient_id + "/Perturbed_TV_PertPoint_" + str(point[0]) + "_" + str(
                                point[1]) + "_" + str(point[2]) + ".nii.gz",
                        )

                    perturb_point = np.zeros_like(gt_prediction)

                    # max/mean value of oar written into perturb location
                    for oar in list_oar_names:
                        # get prediction (pert, gt) on only the oar
                        temp_pred_gt_oar = np.multiply(gt_prediction, dict_images[oar])
                        temp_pred_pert_oar = np.multiply(prediction, dict_images[oar])

                        temp_pert_oar_pt = temp_pred_pert_oar

                        size_oar = np.count_nonzero(temp_pred_gt_oar)

                        # calculate values of interest of the OAR
                        max_gt_oar = np.max(temp_pred_gt_oar)
                        max_pert_oar = np.max(temp_pred_pert_oar)
                        mean_gt_oar = np.divide(np.sum(temp_pred_gt_oar), size_oar)
                        mean_pert_oar = np.divide(np.sum(temp_pred_pert_oar), size_oar)
                        absdiff_oar = np.sum(abs(temp_pred_gt_oar - temp_pred_pert_oar))
                        deltamax_oar = np.abs(max_gt_oar - max_pert_oar)
                        deltamean_oar = np.abs(mean_gt_oar - mean_pert_oar)

                        # Save values of interest
                        # perturb_prediction[oar][point[0], point[1], point[2]] = max_pert_oar
                        perturb_prediction_max[oar][point[0], point[1], point[2]] = max_pert_oar
                        perturb_prediction_mean[oar][point[0], point[1], point[2]] = mean_pert_oar
                        perturb_prediction_dmax[oar][point[0], point[1], point[2]] = deltamax_oar
                        perturb_prediction_dmean[oar][point[0], point[1], point[2]] = deltamean_oar



                        # Location of max 20 points

                        ind = {}
                        temp_max = {}

                        for i in range(0, 20, 1):
                            temp_max[i] = np.max(temp_pert_oar_pt)
                            ind[i] = np.argwhere(temp_pred_pert_oar == temp_max[i])
                            temp_pert_oar_pt = np.where(temp_pert_oar_pt == temp_max[i], 0, temp_pert_oar_pt)


                        for i in range(0, 20, 1):
                            perturb_pointcorrMax[oar][ind[i][0][1], ind[i][0][2], ind[i][0][3]] = 1


                        data_toAdd = [oar, point[0], point[1], point[2], ind[0][0][1], ind[0][0][2], ind[0][0][3],
                                      temp_max[0], ind[1][0][1], ind[1][0][2], ind[1][0][3], temp_max[1], ind[2][0][1],
                                      ind[2][0][2], ind[2][0][3], temp_max[2], ind[3][0][1], ind[3][0][2], ind[3][0][3],
                                      temp_max[2], ind[4][0][1], ind[4][0][2], ind[4][0][3], temp_max[4], ind[5][0][1],
                                      ind[5][0][2], ind[5][0][3], temp_max[5], ind[6][0][1], ind[6][0][2], ind[6][0][3],
                                      temp_max[6], ind[7][0][1], ind[7][0][2], ind[7][0][3], temp_max[7], ind[8][0][1],
                                      ind[8][0][2], ind[8][0][3], temp_max[8], ind[9][0][1], ind[9][0][2], ind[9][0][3],
                                      temp_max[9], ind[10][0][1], ind[10][0][2], ind[10][0][3], temp_max[10],
                                      ind[11][0][1], ind[11][0][2], ind[11][0][3], temp_max[11], ind[12][0][1],
                                      ind[12][0][2], ind[12][0][3], temp_max[12], ind[13][0][1], ind[13][0][2],
                                      ind[13][0][3], temp_max[13], ind[14][0][1], ind[14][0][2], ind[14][0][3],
                                      temp_max[14], ind[15][0][1], ind[15][0][2], ind[15][0][3], temp_max[15],
                                      ind[16][0][1], ind[16][0][2], ind[16][0][3], temp_max[16], ind[17][0][1],
                                      ind[17][0][2], ind[17][0][3], temp_max[17], ind[18][0][1], ind[18][0][2],
                                      ind[18][0][3], temp_max[18], ind[19][0][1], ind[19][0][2], ind[19][0][3],
                                      temp_max[19]]

                        sheet.append(data_toAdd)


                    # get prediction (pert, gt) on only the tv
                    temp_pred_gt = np.multiply(gt_prediction, og_tv)
                    temp_pred_pert = np.multiply(prediction, dict_images[organ])
                    size_tv_gt = np.count_nonzero(temp_pred_gt)
                    size_tv_pert = np.count_nonzero(temp_pred_pert)

                    temp_pert_tv_pt = temp_pred_pert

                    # calculate values of interest of the TV
                    max_gt_tv = np.max(temp_pred_gt)
                    max_pert_tv = np.max(temp_pred_pert)
                    mean_gt_tv = np.divide(np.sum(temp_pred_gt), size_tv_gt)
                    mean_pert_tv = np.divide(np.sum(temp_pred_pert), size_tv_pert)
                    absdiff = np.sum(abs(temp_pred_gt - temp_pred_pert))
                    deltamax_tv  = np.abs(max_gt_tv - max_pert_tv)
                    deltamean_tv = np.abs(mean_gt_tv - mean_pert_tv)

                    ind = {}
                    temp_max = {}

                    for i in range(0, 20, 1):
                        temp_max[i] = np.max(temp_pert_tv_pt)
                        ind[i] = np.argwhere(temp_pred_pert == temp_max[i])
                        temp_pert_tv_pt = np.where(temp_pert_tv_pt == temp_max[i], 0, temp_pert_tv_pt)


                    data_toAdd = [organ, point[0], point[1], point[2], ind[0][0][1], ind[0][0][2], ind[0][0][3],
                                  temp_max[0], ind[1][0][1], ind[1][0][2], ind[1][0][3], temp_max[1], ind[2][0][1],
                                  ind[2][0][2], ind[2][0][3], temp_max[2], ind[3][0][1], ind[3][0][2], ind[3][0][3],
                                  temp_max[2], ind[4][0][1], ind[4][0][2], ind[4][0][3], temp_max[4], ind[5][0][1],
                                  ind[5][0][2], ind[5][0][3], temp_max[5], ind[6][0][1], ind[6][0][2], ind[6][0][3],
                                  temp_max[6], ind[7][0][1], ind[7][0][2], ind[7][0][3], temp_max[7], ind[8][0][1],
                                  ind[8][0][2], ind[8][0][3], temp_max[8], ind[9][0][1], ind[9][0][2], ind[9][0][3],
                                  temp_max[9], ind[10][0][1], ind[10][0][2], ind[10][0][3], temp_max[10],
                                  ind[11][0][1], ind[11][0][2], ind[11][0][3], temp_max[11], ind[12][0][1],
                                  ind[12][0][2], ind[12][0][3], temp_max[12], ind[13][0][1], ind[13][0][2],
                                  ind[13][0][3], temp_max[13], ind[14][0][1], ind[14][0][2], ind[14][0][3],
                                  temp_max[14], ind[15][0][1], ind[15][0][2], ind[15][0][3], temp_max[15],
                                  ind[16][0][1], ind[16][0][2], ind[16][0][3], temp_max[16], ind[17][0][1],
                                  ind[17][0][2], ind[17][0][3], temp_max[17], ind[18][0][1], ind[18][0][2],
                                  ind[18][0][3], temp_max[18], ind[19][0][1], ind[19][0][2], ind[19][0][3],
                                  temp_max[19]]

                    sheet.append(data_toAdd)
                    wb.save(nameFile)

                    # save values of interenst
                    # perturb_prediction[organ][point[0], point[1], point[2]] = max_pert_tv
                    perturb_prediction_max[organ][point[0], point[1], point[2]] = max_pert_tv
                    perturb_prediction_mean[organ][point[0], point[1], point[2]] = mean_pert_tv
                    perturb_prediction_dmax[organ][point[0], point[1], point[2]] = deltamax_tv
                    perturb_prediction_dmean[organ][point[0], point[1], point[2]] = deltamean_tv


                    # Output of nii files
                    for oar in list_oar_names:
                        templete_nii = sitk.ReadImage(patient_dir + "/Dose_Mask.nii.gz")


                        prediction_pointCorrMax_nii = sitk.GetImageFromArray(perturb_pointcorrMax[oar])

                        if sys == 'Windows':
                            os.makedirs(save_path + "\\" + patient_id, exist_ok=True)
                            sitk.WriteImage(
                                prediction_pointCorrMax_nii,
                                save_path + "\\" + patient_id + "/Perturbed_TV_PointCloud_" + str(point[0]) + "_" + str(point[1]) + "_" + str(point[2]) + "_" + oar + ".nii.gz",
                            )
                        else:
                            os.makedirs(save_path + "/" + patient_id, exist_ok=True)
                            sitk.WriteImage(
                                prediction_pointCorrMax_nii,
                                save_path + "/" + patient_id + "/Perturbed_TV_PointCloud_" + str(point[0]) + "_" + str(point[1]) + "_" + str(point[2]) + "_" + oar + ".nii.gz",
                            )

                templete_nii = sitk.ReadImage(patient_dir + "/Dose_Mask.nii.gz")



if __name__ == "__main__":

    root_dir = "/Users/zahir/Documents/Github/astra/"
    # root_dir = "/home/studentshare/Documents/astra/"
    # root_dir = "/storage/homefs/zm13j051/astra/"
    # root_dir = os.getcwd()
    model_dir = os.path.join(root_dir, "models")
    output_dir = os.path.join(root_dir, "output_perturb")
    os.makedirs(output_dir, exist_ok=True)

    gt_dir = os.path.join(root_dir, "data", "processed-dldp")
    test_dir = gt_dir  # change this if somewhere else.

    if not os.path.exists(model_dir):
        raise Exception(
            "OpenKBP_C3D should be prepared before testing, please run prepare_OpenKBP_C3D.py"
        )

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--GPU_id", type=int, default=-1, help="GPU id used for testing (default: 0)"
    )
    parser.add_argument(
        "--model_path",
        type=str,
        default=os.path.join(model_dir, "best_val_evaluation_index.pkl"),
    )
    parser.add_argument(
        "--TTA", type=bool, default=True, help="do test-time augmentation, default True"
    )
    args = parser.parse_args()

    trainer_ = NetworkTrainer()
    trainer_.setting.project_name = "C3D"
    trainer_.setting.output_dir = output_dir

    trainer_.setting.network = Model(
        in_ch=15,
        out_ch=1,
        list_ch_A=[-1, 16, 32, 64, 128, 256],
        list_ch_B=[-1, 32, 64, 128, 256, 512],
    )

    # Load model weights
    trainer_.init_trainer(
        ckpt_file=args.model_path, list_GPU_ids=[args.GPU_id], only_network=True
    )

    # for subject_id in [90, 82, 81, 88]:
    # after training rerun 70-100 are allowed without 77
    # 84, 73, 98
    for subject_id in [98]:
        # Start inference
        print("\n\n# Start inference !")
        list_patient_dirs = [os.path.join(test_dir, "DLDP_" + str(subject_id).zfill(3))]
        inference_with_perturbation(
            trainer_,
            list_patient_dirs,
            save_path=os.path.join(trainer_.setting.output_dir, "Prediction_" + PERT_TYPE + str(PERT_SIZE)),
            do_TTA=args.TTA,
        )
